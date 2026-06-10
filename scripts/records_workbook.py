#!/usr/bin/env python3
"""Shared V&V Records Workbook import/export helpers."""

from __future__ import annotations

import datetime as dt
import subprocess
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[1]
RECORDS_DIR = ROOT / "records"
WORKBOOK_SCHEMA_VERSION = "1"
DEFAULT_WORKBOOK = ROOT / "vnv-records.xlsx"

LIST_SEPARATOR = "\n"


@dataclass(frozen=True)
class SheetSpec:
    name: str
    yaml_file: str
    yaml_key: str
    columns: tuple[str, ...]
    list_columns: frozenset[str] = frozenset()
    integer_columns: frozenset[str] = frozenset()
    required_columns: frozenset[str] = frozenset({"id"})
    single_row: bool = False


SHEETS: tuple[SheetSpec, ...] = (
    SheetSpec(
        name="Project",
        yaml_file="project.yaml",
        yaml_key="project",
        columns=(
            "name",
            "product_code",
            "version",
            "manufacturer.name",
            "intended_use",
            "intended_users",
            "operating_environment",
            "software_safety_class",
            "regulatory_context.primary_standard",
            "regulatory_context.lifecycle_scope",
            "limitations",
        ),
        list_columns=frozenset({"intended_users", "operating_environment", "limitations"}),
        required_columns=frozenset({"name", "product_code", "version"}),
        single_row=True,
    ),
    SheetSpec(
        name="Documents",
        yaml_file="documents.yaml",
        yaml_key="documents",
        columns=("id", "number", "title", "version", "status", "output"),
        required_columns=frozenset({"id", "number", "title", "version", "status", "output"}),
    ),
    SheetSpec(
        name="Software Development Plan",
        yaml_file="software-development-plan.yaml",
        yaml_key="plan",
        columns=(
            "lifecycle_model",
            "configuration_management",
            "problem_resolution",
            "verification_strategy",
            "ai_model_controls",
        ),
        list_columns=frozenset({"ai_model_controls"}),
        required_columns=frozenset({"lifecycle_model", "verification_strategy"}),
        single_row=True,
    ),
    SheetSpec(
        name="Requirements",
        yaml_file="requirements.yaml",
        yaml_key="requirements",
        columns=(
            "id",
            "title",
            "description",
            "rationale",
            "related_architecture",
            "related_design",
            "risk_controls",
            "verified_by",
        ),
        list_columns=frozenset(
            {"related_architecture", "related_design", "risk_controls", "verified_by"}
        ),
        required_columns=frozenset({"id", "title", "description", "rationale", "verified_by"}),
    ),
    SheetSpec(
        name="Architecture",
        yaml_file="architecture.yaml",
        yaml_key="architecture_items",
        columns=(
            "id",
            "title",
            "description",
            "related_requirements",
            "related_design",
            "verified_by",
        ),
        list_columns=frozenset({"related_requirements", "related_design", "verified_by"}),
        required_columns=frozenset({"id", "title", "description"}),
    ),
    SheetSpec(
        name="Detailed Design",
        yaml_file="detailed-design.yaml",
        yaml_key="design_items",
        columns=("id", "title", "description", "related_architecture", "verified_by"),
        list_columns=frozenset({"related_architecture", "verified_by"}),
        required_columns=frozenset({"id", "title", "description"}),
    ),
    SheetSpec(
        name="Unit Tests",
        yaml_file="tests.yaml",
        yaml_key="unit_tests",
        columns=("id", "title", "verifies", "procedure", "acceptance_criteria"),
        list_columns=frozenset({"verifies"}),
        required_columns=frozenset({"id", "title", "verifies", "procedure", "acceptance_criteria"}),
    ),
    SheetSpec(
        name="Integration Tests",
        yaml_file="tests.yaml",
        yaml_key="integration_tests",
        columns=("id", "title", "verifies", "procedure", "acceptance_criteria"),
        list_columns=frozenset({"verifies"}),
        required_columns=frozenset({"id", "title", "verifies", "procedure", "acceptance_criteria"}),
    ),
    SheetSpec(
        name="System Tests",
        yaml_file="tests.yaml",
        yaml_key="system_tests",
        columns=("id", "title", "verifies", "procedure", "acceptance_criteria"),
        list_columns=frozenset({"verifies"}),
        required_columns=frozenset({"id", "title", "verifies", "procedure", "acceptance_criteria"}),
    ),
    SheetSpec(
        name="Risk Controls",
        yaml_file="risk-controls.yaml",
        yaml_key="risk_controls",
        columns=("id", "title", "risk", "control", "related_requirements", "verified_by"),
        list_columns=frozenset({"related_requirements", "verified_by"}),
        required_columns=frozenset(
            {"id", "title", "risk", "control", "related_requirements", "verified_by"}
        ),
    ),
    SheetSpec(
        name="AI Models",
        yaml_file="ai-models.yaml",
        yaml_key="ai_models",
        columns=(
            "id",
            "name",
            "version",
            "task",
            "input.modality",
            "input.format",
            "output.type",
            "output.target",
            "intended_use_limitation",
            "related_requirements",
            "related_architecture",
            "related_design",
            "related_system_tests",
        ),
        list_columns=frozenset(
            {
                "related_requirements",
                "related_architecture",
                "related_design",
                "related_system_tests",
            }
        ),
        required_columns=frozenset(
            {"id", "name", "version", "task", "related_requirements", "related_system_tests"}
        ),
    ),
    SheetSpec(
        name="Datasets",
        yaml_file="datasets.yaml",
        yaml_key="datasets",
        columns=(
            "id",
            "name",
            "purpose",
            "modality",
            "sample_count",
            "anatomy",
            "inclusion_criteria",
            "exclusion_criteria",
            "related_model",
            "related_system_tests",
        ),
        list_columns=frozenset(
            {"inclusion_criteria", "exclusion_criteria", "related_system_tests"}
        ),
        integer_columns=frozenset({"sample_count"}),
        required_columns=frozenset(
            {"id", "name", "purpose", "sample_count", "related_model", "related_system_tests"}
        ),
    ),
    SheetSpec(
        name="Performance Metrics",
        yaml_file="performance-metrics.yaml",
        yaml_key="performance_metrics",
        columns=(
            "id",
            "name",
            "description",
            "acceptance_criterion",
            "related_model",
            "related_system_tests",
        ),
        list_columns=frozenset({"related_system_tests"}),
        required_columns=frozenset(
            {"id", "name", "acceptance_criterion", "related_model", "related_system_tests"}
        ),
    ),
)

CHILD_SHEET_COLUMNS = {
    "Document Approvers": ("document_id", "order", "role", "name"),
    "Revision History": ("version", "date", "description", "author"),
}


def load_records(records_dir: Path = RECORDS_DIR) -> dict[str, Any]:
    records: dict[str, Any] = {}
    for path in sorted(records_dir.glob("*.yaml")):
        with path.open("r", encoding="utf-8") as handle:
            records[path.name] = yaml.safe_load(handle) or {}
    return records


def dump_records(records: dict[str, Any], output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    for filename, data in records.items():
        with (output_dir / filename).open("w", encoding="utf-8") as handle:
            yaml.safe_dump(data, handle, sort_keys=False, allow_unicode=True, width=100)


def get_nested(row: dict[str, Any], dotted_key: str) -> Any:
    value: Any = row
    for part in dotted_key.split("."):
        if not isinstance(value, dict):
            return None
        value = value.get(part)
    return value


def set_nested(row: dict[str, Any], dotted_key: str, value: Any) -> None:
    parts = dotted_key.split(".")
    target = row
    for part in parts[:-1]:
        target = target.setdefault(part, {})
    target[parts[-1]] = value


def format_cell(value: Any, is_list: bool) -> Any:
    if value is None:
        return None
    if is_list:
        values = value if isinstance(value, list) else [value]
        return LIST_SEPARATOR.join(str(item) for item in values)
    return value


def parse_cell(value: Any, column: str, spec: SheetSpec) -> Any:
    if value is None:
        return [] if column in spec.list_columns else None
    if isinstance(value, str) and value.strip() == "":
        return [] if column in spec.list_columns else None
    if column in spec.list_columns:
        if isinstance(value, str):
            return [item.strip() for item in value.replace(",", "\n").splitlines() if item.strip()]
        return [value]
    if column in spec.integer_columns:
        return int(value)
    return value


def add_table_sheet(wb: Workbook, spec: SheetSpec, records: dict[str, Any]) -> None:
    ws = wb.create_sheet(spec.name)
    ws.append(list(spec.columns))
    source = records[spec.yaml_file][spec.yaml_key]
    rows = [source] if spec.single_row else source
    for row in rows:
        ws.append(
            [
                format_cell(get_nested(row, column), column in spec.list_columns)
                for column in spec.columns
            ]
        )
    style_sheet(ws, required_columns=spec.required_columns, list_columns=spec.list_columns)


def add_child_sheets(wb: Workbook, records: dict[str, Any]) -> None:
    approvers = wb.create_sheet("Document Approvers")
    approvers.append(list(CHILD_SHEET_COLUMNS["Document Approvers"]))
    for document in records["documents.yaml"]["documents"]:
        for index, approver in enumerate(document.get("approvers", []), start=1):
            approvers.append([document["id"], index, approver.get("role"), approver.get("name")])
    style_sheet(
        approvers,
        required_columns=frozenset({"document_id", "role", "name"}),
        list_columns=frozenset(),
    )

    revisions = wb.create_sheet("Revision History")
    revisions.append(list(CHILD_SHEET_COLUMNS["Revision History"]))
    for revision in records["documents.yaml"].get("revision_history", []):
        revisions.append(
            [revision.get(column) for column in CHILD_SHEET_COLUMNS["Revision History"]]
        )
    style_sheet(
        revisions,
        required_columns=frozenset({"version", "date", "description", "author"}),
        list_columns=frozenset(),
    )


def style_sheet(ws: Any, required_columns: frozenset[str], list_columns: frozenset[str]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    required_fill = PatternFill("solid", fgColor="FFF2CC")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for index, cell in enumerate(ws[1], start=1):
        column = str(cell.value)
        width = 18
        if column in {
            "description",
            "rationale",
            "procedure",
            "acceptance_criteria",
            "intended_use_limitation",
        }:
            width = 54
        elif column in list_columns:
            width = 28
        ws.column_dimensions[get_column_letter(index)].width = width
        if column in required_columns:
            for row in ws.iter_rows(min_row=2, min_col=index, max_col=index):
                row[0].fill = required_fill
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_readme(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "README"
    rows = [
        ("V&V Records Workbook", "Human-editable Excel view of records/*.yaml."),
        (
            "Canonical source",
            "records/*.yaml remains canonical. Import validates before replacing YAML.",
        ),
        ("List fields", "Enter one ID or value per line inside the cell."),
        ("Hidden sheets", "_meta, _schema, and _lookups support import/export and validation."),
        ("Reserved file name", "vnv-records.xlsx"),
    ]
    for row in rows:
        ws.append(row)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 92
    for cell in ws["A"]:
        cell.font = Font(bold=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_meta(wb: Workbook, source_commit: str | None) -> None:
    ws = wb.create_sheet("_meta")
    ws.append(("key", "value"))
    rows = {
        "workbook_type": "V&V Records Workbook",
        "workbook_schema_version": WORKBOOK_SCHEMA_VERSION,
        "exported_at": dt.datetime.now(dt.UTC).replace(microsecond=0).isoformat(),
        "source_commit": source_commit or "",
        "records_dir": "records",
    }
    for key, value in rows.items():
        ws.append((key, value))
    ws.sheet_state = "hidden"


def add_schema(wb: Workbook) -> None:
    ws = wb.create_sheet("_schema")
    ws.append(("sheet", "yaml_file", "yaml_key", "column", "kind", "required"))
    for spec in SHEETS:
        for column in spec.columns:
            if column in spec.list_columns:
                kind = "list"
            elif column in spec.integer_columns:
                kind = "integer"
            else:
                kind = "scalar"
            ws.append(
                (
                    spec.name,
                    spec.yaml_file,
                    spec.yaml_key,
                    column,
                    kind,
                    column in spec.required_columns,
                )
            )
    ws.sheet_state = "hidden"


def add_lookups(wb: Workbook, records: dict[str, Any]) -> None:
    ws = wb.create_sheet("_lookups")
    items: dict[str, list[str]] = {
        "Record Items": [],
        "Requirements": [],
        "Architecture": [],
        "Design": [],
        "Unit Tests": [],
        "Integration Tests": [],
        "System Tests": [],
        "Risk Controls": [],
        "Models": [],
    }
    for spec in SHEETS:
        source = records[spec.yaml_file][spec.yaml_key]
        rows = [source] if spec.single_row else source
        for row in rows:
            item_id = row.get("id") if isinstance(row, dict) else None
            if not item_id:
                continue
            if item_id.startswith(("SR-", "SA-", "SD-", "UT-", "IT-", "ST-")):
                items["Record Items"].append(item_id)
            if item_id.startswith("SR-"):
                items["Requirements"].append(item_id)
            elif item_id.startswith("SA-"):
                items["Architecture"].append(item_id)
            elif item_id.startswith("SD-"):
                items["Design"].append(item_id)
            elif item_id.startswith("UT-"):
                items["Unit Tests"].append(item_id)
            elif item_id.startswith("IT-"):
                items["Integration Tests"].append(item_id)
            elif item_id.startswith("ST-"):
                items["System Tests"].append(item_id)
            elif item_id.startswith("RC-"):
                items["Risk Controls"].append(item_id)
            elif item_id.startswith("MODEL-"):
                items["Models"].append(item_id)
    ws.append(tuple(items))
    max_len = max(len(values) for values in items.values())
    for index in range(max_len):
        ws.append(
            tuple(values[index] if index < len(values) else None for values in items.values())
        )
    ws.sheet_state = "hidden"


def add_data_validations(wb: Workbook) -> None:
    lookup_map = {
        "related_requirements": "_lookups!$B$2:$B$200",
        "related_architecture": "_lookups!$C$2:$C$200",
        "related_design": "_lookups!$D$2:$D$200",
        "verified_by": "_lookups!$A$2:$A$200",
        "verifies": "_lookups!$A$2:$A$200",
        "risk_controls": "_lookups!$H$2:$H$200",
        "related_model": "_lookups!$I$2:$I$200",
        "related_system_tests": "_lookups!$G$2:$G$200",
    }
    for spec in SHEETS:
        ws = wb[spec.name]
        for index, column in enumerate(spec.columns, start=1):
            if column not in lookup_map:
                continue
            letter = get_column_letter(index)
            validation = DataValidation(type="list", formula1=lookup_map[column], allow_blank=True)
            validation.error = "Choose a known ID from the V&V Records lookup list."
            validation.errorTitle = "Unknown Record Item"
            ws.add_data_validation(validation)
            validation.add(f"{letter}2:{letter}500")


def source_commit() -> str | None:
    try:
        result = subprocess.run(
            ["git", "rev-parse", "--short", "HEAD"],
            cwd=ROOT,
            check=True,
            text=True,
            capture_output=True,
        )
    except Exception:
        return None
    return result.stdout.strip()


def export_workbook(records_dir: Path, output_path: Path) -> None:
    records = load_records(records_dir)
    wb = Workbook()
    add_readme(wb)
    for spec in SHEETS:
        add_table_sheet(wb, spec, records)
    add_child_sheets(wb, records)
    add_meta(wb, source_commit())
    add_schema(wb)
    add_lookups(wb, records)
    add_data_validations(wb)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def read_table(ws: Any, required_columns: frozenset[str]) -> list[dict[str, Any]]:
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in ws[1]]
    missing = sorted(required_columns - set(headers))
    if missing:
        raise ValueError(f"{ws.title} missing required columns: {', '.join(missing)}")
    rows: list[dict[str, Any]] = []
    for cells in ws.iter_rows(min_row=2, values_only=True):
        if not any(value not in (None, "") for value in cells):
            continue
        rows.append(
            {header: value for header, value in zip(headers, cells, strict=False) if header}
        )
    return rows


def import_workbook(workbook_path: Path) -> dict[str, Any]:
    wb = load_workbook(workbook_path, data_only=False)
    check_workbook_meta(wb)
    records: dict[str, Any] = {}
    for spec in SHEETS:
        if spec.name not in wb.sheetnames:
            raise ValueError(f"Workbook is missing sheet: {spec.name}")
        source_rows = read_table(wb[spec.name], spec.required_columns)
        parsed_rows = []
        for source_row in source_rows:
            row: dict[str, Any] = {}
            for column in spec.columns:
                value = parse_cell(source_row.get(column), column, spec)
                if value is None or value == []:
                    continue
                set_nested(row, column, value)
            parsed_rows.append(row)
        records.setdefault(spec.yaml_file, {})[spec.yaml_key] = (
            parsed_rows[0] if spec.single_row else parsed_rows
        )
    attach_document_children(wb, records)
    return records


def check_workbook_meta(wb: Any) -> None:
    if "_meta" not in wb.sheetnames:
        raise ValueError("Workbook is missing _meta sheet.")
    values = {
        str(row[0]).strip(): str(row[1]).strip()
        for row in wb["_meta"].iter_rows(min_row=2, values_only=True)
        if row[0] is not None
    }
    if values.get("workbook_type") != "V&V Records Workbook":
        raise ValueError("Workbook is not a V&V Records Workbook.")
    if values.get("workbook_schema_version") != WORKBOOK_SCHEMA_VERSION:
        raise ValueError(
            "Unsupported V&V Records Workbook schema version: "
            f"{values.get('workbook_schema_version')}"
        )


def attach_document_children(wb: Any, records: dict[str, Any]) -> None:
    if "Document Approvers" not in wb.sheetnames:
        raise ValueError("Workbook is missing sheet: Document Approvers")
    if "Revision History" not in wb.sheetnames:
        raise ValueError("Workbook is missing sheet: Revision History")

    documents = {item["id"]: item for item in records["documents.yaml"]["documents"]}
    for row in read_table(wb["Document Approvers"], frozenset({"document_id", "role", "name"})):
        document_id = str(row.get("document_id", "")).strip()
        if document_id not in documents:
            raise ValueError(f"Document Approvers references unknown document: {document_id}")
        documents[document_id].setdefault("approvers", []).append(
            {"role": row.get("role"), "name": row.get("name")}
        )

    revisions = []
    for row in read_table(
        wb["Revision History"], frozenset({"version", "date", "description", "author"})
    ):
        revisions.append(
            {column: row.get(column) for column in CHILD_SHEET_COLUMNS["Revision History"]}
        )
    records["documents.yaml"]["revision_history"] = revisions


def validate_records_dir(records_dir: Path) -> None:
    subprocess.run(
        ["uv", "run", "python", "scripts/validate_records.py", "--records-dir", str(records_dir)],
        cwd=ROOT,
        check=True,
    )
